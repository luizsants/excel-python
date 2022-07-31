from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import *

wb = load_workbook('relatorio.xlsx')
ws = wb.active

for i in range(7):
    aux = 2
    ws.delete_cols(aux)

for i in range(7):
    aux = 4
    ws.delete_cols(aux)

for i in range(2):
    aux = 5
    ws.delete_cols(aux)

for i in range(11):
    aux = 6
    ws.delete_cols(aux)

for i in range(4):
    aux = 7
    ws.delete_cols(aux)

for col in range(1, 7): 
    ws[get_column_letter(col) + '1'].font = Font(bold = True)
    col_letter = get_column_letter(col)

#Select best width to the columns
for i in range(1, ws.max_row+1):
    ws.row_dimensions[i].height = 15

for i in range(1, ws.max_column+1):
    all_rows = []
    all_rows.append(0)

    if (get_column_letter(i) == 'B'):
        for j in range(1, ws.max_row+1):
            all_rows.append(len(ws['B'+str(j)].value))
            if (all_rows[j] > all_rows[j-1]):
                max_verified = all_rows[j]
        ws.column_dimensions[get_column_letter(i)].width = max_verified
    
    elif(get_column_letter(i) == 'C'):
        for j in range(1, ws.max_row+1):
            all_rows.append(len(ws['C'+str(j)].value))
            if (all_rows[j] > all_rows[j-1]):
                max_verified = all_rows[j]
        ws.column_dimensions[get_column_letter(i)].width = max_verified

    elif(get_column_letter(i) == 'D'):
        for j in range(1, ws.max_row+1):
            all_rows.append(len(ws['D'+str(j)].value))
            if (all_rows[j] > all_rows[j-1]):
                max_verified = all_rows[j]
        ws.column_dimensions[get_column_letter(i)].width = max_verified

    elif(get_column_letter(i) == 'E'):
        ws.column_dimensions[get_column_letter(i)].width = 35

    else:  
        for j in range(1, ws.max_row+1):
            all_rows.append(len(ws['F'+str(j)].value))
            if (all_rows[j] > all_rows[j-1]):
                max_verified = all_rows[j]
        ws.column_dimensions[get_column_letter(i)].width = max_verified

wb.save('relatorio1.xlsx')